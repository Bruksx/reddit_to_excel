from exe import Fetch
from openpyxl import Workbook, load_workbook
from bot import dataBot

class Sheetx:
    def __init__(self):
        try:
            self.wb = load_workbook('spreadsheets/test.xlsx')
            self.ws = self.wb.active
        except FileNotFoundError:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.wb.save('spreadsheets/test.xlsx')
    
    def insert_subgroup_data(self, subgroup):
        subreddits = Fetch().get_subreddits_in_subgroup(subgroup)
        sheet = self.wb.create_sheet(subgroup)
        self.creat_table(subgroup)
        for i in subreddits:
            data = self.select(i)
            try:
                pos = len(list(sheet.iter_rows())) + 1
                j = 1
                for key in data:
                    sheet.cell(row = pos, column = j, value = key)
                    j += 1
            except TypeError:
                pass
        self.wb.save("spreadsheets/test.xlsx")


    def creat_table(self, ws):
        sheet = self.wb[ws]
        li = [
            "Subreddit",
            "All_mods",
            "Number of subscribers",
            "Text type",
            "Non text type",
            "Number of mods",
            "List of mods",
            "Number of moderator posts",
            "Moderator to subscribers ratio",
            "Average comment score",
            "Number of post",
            "Total comments",
            "Average comment per post",
            "List of flairs",
            "Number of automod post",
            "List of automod flair",
            ]

        for i in range(len(li)):
            sheet.cell(row=1, column= i+1, value= li[i])
            self.wb.save("spreadsheets/test.xlsx")
    
    def select(self, subreddit):
        for i in range(1, 23):
            val = self.ws.cell(row = i, column = 1)
            if val.value == subreddit:
                li = [subreddit]
                end = len(list(self.ws.iter_rows()))
                for j in range(2, end):
                    new_vals = self.ws.cell(row=i, column= j)
                    li.append(new_vals.value)
                return li
    
    def insert_into_table(self, sub):
        self.wb = load_workbook('spreadsheets/test.xlsx')
        self.ws = self.wb.active
        data = Fetch().all_data(sub)
        pos = len(list(self.ws.iter_rows())) + 1
        i = 1
        for key in data:
            self.ws.cell(row = pos, column = i, value = data[key])
            i += 1
        self.wb.save("spreadsheets/test.xlsx")
    
    def insert_all(self):
        all_subs = dataBot.get_list_of_subreddits()
        for i in all_subs:
            print(f"fetching {i}")
            self.insert_into_table(i)
        

s = Sheetx()

s.creat_table("Sheet")
s.insert_all()
s.insert_subgroup_data("hiphop")